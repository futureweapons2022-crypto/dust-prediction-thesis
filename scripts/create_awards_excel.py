"""Create Excel sheet for Studies & Business Development section awards 2026."""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Awards 2026"
ws.sheet_view.rightToLeft = True

# Styles
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
urgent_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
verified_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
unverified_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
wrap = Alignment(wrap_text=True, vertical="top", horizontal="right")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# Title row
ws.merge_cells("A1:P1")
title_cell = ws["A1"]
title_cell.value = "مشاركات قسم الدراسات وتطوير الأعمال — جوائز 2026"
title_cell.font = Font(bold=True, size=14, color="2F5496")
title_cell.alignment = Alignment(horizontal="center", vertical="center")

# Headers
headers = [
    "#",
    "القسم المشارك",
    "اسم الجائزة",
    "نبذة عن الجائزة",
    "فئات الجائزة",
    "تصنيف الجائزة",
    "الموقع الإلكتروني",
    "الجهة المانحة للجائزة",
    "آخر موعد للمشاركة",
    "رسوم الجائزة",
    "الموضوع / المحور",
    "المشاريع المدمجة من الأقسام",
    "الأقسام المساهمة",
    "حالة المشاركة",
    "حالة التحقق",
    "الملاحظات",
]

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    cell.border = thin_border

# === 6 PARTICIPATIONS — VERIFIED STATUS INCLUDED ===

participations = [
    {
        "num": 1,
        "section": "قسم الدراسات وتطوير الأعمال",
        "award": "I2SL Sustainable Labs Award 2026",
        "about": (
            "جائزة دولية من المعهد الدولي للمختبرات المستدامة (I2SL) تكرّم المشاريع "
            "والبرامج والأشخاص الذين حققوا إنجازات مبتكرة في استدامة المختبرات وكفاءة "
            "الطاقة وتقليل الانبعاثات والنفايات. الحفل في مؤتمر بوسطن 15 سبتمبر 2026."
        ),
        "category": "Lab Programs or Initiatives",
        "classification": "B (دولية)",
        "website": "https://www.i2sl.org/sustainable-lab-awards",
        "organizer": "International Institute for Sustainable Laboratories (I2SL)",
        "deadline": "09/03/2026",
        "fees": "غير محدد",
        "topic": (
            "برنامج الاستدامة المتكامل في مختبرات دبي المركزية: "
            "تطوير أساليب فحص أقل استهلاكاً للمواد الكيميائية وتقييم البصمة الكربونية للمنتجات"
        ),
        "projects": (
            "1. تحسين كشف الفوسفات بتقنية IC بدل الطيف الضوئي — تقليل النفايات الكيميائية (الكيميائي)\n"
            "2. تقييم استدامة حفاضات الأطفال والبصمة الكربونية (الكيميائي)\n"
            "3. شهادات المطابقة لمنتجات الأبنية الخضراء (منح الشهادات)"
        ),
        "contributing": "الكيميائي، منح الشهادات",
        "status": "عاجل — الموعد النهائي 9 مارس (11 يوم)",
        "verified": "تم التحقق — التقديم مفتوح، 3-5 صفحات PDF، ترسل إلى info@i2sl.org",
        "notes": (
            "جائزة جديدة غير موجودة في السجل الحالي. "
            "مخصصة للمختبرات تحديداً. "
            "المشاريع يجب أن تكون مكتملة وتعمل منذ سنة على الأقل."
        ),
    },
    {
        "num": 2,
        "section": "قسم الدراسات وتطوير الأعمال",
        "award": "OECD Call for Government Innovations (8th Edition)",
        "about": (
            "دعوة مفتوحة من منظمة التعاون الاقتصادي والتنمية لجميع الحكومات في العالم "
            "لمشاركة ابتكاراتها. يتم نشر المشاركات المقبولة في قاعدة بيانات OECD العالمية. "
            "مفتوحة لجميع الدول — أي جغرافيا، أي مستوى حكومي."
        ),
        "category": "Simplifying Government (تبسيط العمليات الحكومية)",
        "classification": "A (دولية)",
        "website": "https://oecd-opsi.org/blog/the-oecd-call-for-government-innovations-2026-edition/",
        "organizer": "OECD Observatory of Public Sector Innovation (OPSI)",
        "deadline": "31/03/2026",
        "fees": "مجاني",
        "topic": (
            "التحول الرقمي والذكاء الاصطناعي الشامل في مختبرات دبي المركزية: "
            "استراتيجية تنسيقية لنشر الذكاء الاصطناعي عبر 4 مختبرات متخصصة"
        ),
        "projects": (
            "1. الروبوت الذكي لفحص الإسمنت (البنية التحتية)\n"
            "2. نظام FCT لفحص أدوات تنظيف الأرضيات بالذكاء الاصطناعي (الكهروميكانيكية)\n"
            "3. برنامج VeriSoft لرقمنة التحقق القانوني (المقاييس)\n"
            "4. المختبر المتنقل الذكي (الميكروبيولوجي)\n"
            "5. مختبر دبي الذكي (الميكروبيولوجي)"
        ),
        "contributing": "البنية التحتية، الكهروميكانيكية، المقاييس، الميكروبيولوجي",
        "status": "قيد الإعداد — الموعد النهائي 31 مارس",
        "verified": (
            "تم التحقق — الصفحة نشرت 24 فبراير 2026 على موقع OECD الرسمي. "
            "الموعد النهائي 31 مارس مؤكد. نموذج التقديم متاح أونلاين."
        ),
        "notes": (
            "جائزة جديدة غير موجودة في السجل الحالي. مجانية. "
            "المعايير: الابتكار يجب أن يكون جديداً + مطبقاً + له أثر. "
            "أولويات 2026: تبسيط العمليات الحكومية، التفاعل مع المواطنين، قياس الأداء."
        ),
    },
    {
        "num": 3,
        "section": "قسم الدراسات وتطوير الأعمال",
        "award": "PMI Project of the Year Award",
        "about": (
            "جائزة من معهد إدارة المشاريع (PMI) تكرّم المشاريع المتميزة من جميع القطاعات "
            "والصناعات. تُقيّم المشاريع ضمن 3 فئات: التكنولوجيا/الهندسة، الاجتماعية، البناء/البنية التحتية. "
            "لا تتطلب عضوية PMI."
        ),
        "category": "Social Impact",
        "classification": "A (دولية)",
        "website": "https://www.pmi.org/about/awards/professional/project-of-the-year",
        "organizer": "Project Management Institute (PMI)",
        "deadline": "01/04/2026",
        "fees": "مجاني",
        "topic": (
            "مشروع حماية المستهلك عبر المختبرات: إدارة 7 مسارات عمل عبر 3 مختبرات "
            "لبناء منظومة متكاملة للكشف عن المخاطر الصحية الناشئة في دبي"
        ),
        "projects": (
            "1. فحوصات الأغذية المبتكرة — الحشرات والكائنات المعدلة وراثياً (الميكروبيولوجي)\n"
            "2. السموم الطحلبية Phycotoxins (الميكروبيولوجي)\n"
            "3. كشف TPO في طلاء الأظافر + مسح سوق دبي (الكيميائي)\n"
            "4. Bisphenol A في الورق الحراري + نموذج التعرض البشري (الكيميائي)\n"
            "5. خدمات العمر الافتراضي Shelf-Life (الميكروبيولوجي)\n"
            "6. فحص Legionella بتقنية Mica (الميكروبيولوجي)"
        ),
        "contributing": "الميكروبيولوجي، الكيميائي",
        "status": "قيد الإعداد — الموعد النهائي 1 أبريل",
        "verified": (
            "تم التحقق — الموعد 1 أبريل 2026 مؤكد. مجاني. مقال 2000 كلمة. "
            "المشروع يجب أن يكون مكتملاً خلال آخر 18 شهر. "
            "القطاع العام مؤهل صراحة."
        ),
        "notes": (
            "جائزة جديدة غير موجودة في السجل الحالي (PMI موجود في سجل البنية التحتية لكن بفئة مختلفة). "
            "PMI يريد مشروعاً واحداً — يجب تقديم المسارات السبعة كمشروع واحد منسق."
        ),
    },
    {
        "num": 4,
        "section": "قسم الدراسات وتطوير الأعمال",
        "award": "جائزة حميد بن راشد للاستدامة الدولية",
        "about": (
            "تهدف إلى تشجيع المؤسسات على تبني أفضل الممارسات الخضراء للحفاظ على البيئة "
            "والموارد الطبيعية وتحقيق التنمية المستدامة. 4 فئات: أفضل بحث بيئي، "
            "شخصية بيئية، مؤسسات مستدامة، أفكار مستدامة."
        ),
        "category": "المؤسسات المستدامة",
        "classification": "A (إقليمية)",
        "website": "https://aiec.am.gov.ae/hbrisa/",
        "organizer": "دائرة البلدية والتخطيط — عجمان",
        "deadline": "30/04/2026",
        "fees": "غير محدد",
        "topic": (
            "منظومة الاستدامة المتكاملة في مختبرات دبي المركزية: "
            "تقليل النفايات الكيميائية وتقييم البصمة الكربونية للمنتجات الاستهلاكية"
        ),
        "projects": (
            "1. تحسين كشف الفوسفات بتقنية IC بدل الطيف الضوئي التقليدي — تقليل النفايات الكيميائية (الكيميائي)\n"
            "2. تقييم استدامة حفاضات الأطفال والبصمة الكربونية (الكيميائي)\n"
            "3. شهادات المطابقة لمنتجات الأبنية الخضراء (منح الشهادات)"
        ),
        "contributing": "الكيميائي، منح الشهادات",
        "status": "قيد الإعداد — باب التقديم مفتوح",
        "verified": (
            "تم التحقق — رابط التسجيل فعّال: aiec-registration.com/humaid-bin-rashid-award-registration-2026. "
            "الموعد النهائي 30 أبريل مؤكد. القائمة المختصرة يوليو، الفائزون أكتوبر 2026."
        ),
        "notes": (
            "موجودة في السجل الحالي — لم يشارك أي قسم بعد. "
            "المعايير: الابتكار + القابلية للتطبيق. "
            "يجب التركيز على الأثر البيئي (كميات النفايات المخفضة، المواد الكيميائية الملغاة) وليس كفاءة المختبر."
        ),
    },
    {
        "num": 5,
        "section": "قسم الدراسات وتطوير الأعمال",
        "award": "Stevie International Business Awards (23rd Edition)",
        "about": (
            "جوائز ستيفي الدولية للأعمال — من أعرق جوائز الأعمال في العالم. "
            "تكرّم المؤسسات والأفراد من جميع القطاعات والأحجام حول العالم. "
            "جديد 2026: فئات مخصصة للقطاع الحكومي والذكاء الاصطناعي. "
            "حفل التكريم في باريس، 28 أكتوبر 2026. جوائز ذهبية وفضية وبرونزية."
        ),
        "category": "Public Sector & Government Innovation",
        "classification": "A (دولية)",
        "website": "https://iba.stevieawards.com/",
        "organizer": "Stevie Awards Inc.",
        "deadline": "08/04/2026 (Early Bird) — 06/05/2026 (Primary) — 17/06/2026 (Final)",
        "fees": "~$510 (~1,870 درهم) لكل مشاركة",
        "topic": (
            "الابتكار الحكومي المتكامل: كيف نسّقت مختبرات دبي المركزية نشر الذكاء الاصطناعي "
            "والتحول الرقمي عبر 6 مختبرات متخصصة لتحسين الخدمات العامة وحماية المستهلك"
        ),
        "projects": (
            "1. الروبوت الذكي لفحص الإسمنت — أتمتة كاملة بالذكاء الاصطناعي (البنية التحتية)\n"
            "2. نظام FCT لفحص أدوات تنظيف الأرضيات بالذكاء الاصطناعي (الكهروميكانيكية)\n"
            "3. برنامج VeriSoft لرقمنة التحقق القانوني (المقاييس)\n"
            "4. المختبر المتنقل الذكي (الميكروبيولوجي)\n"
            "5. مختبر دبي الذكي (الميكروبيولوجي)\n"
            "6. التحقق المترولوجي المتنقل OTG (المقاييس)"
        ),
        "contributing": "البنية التحتية، الكهروميكانيكية، المقاييس، الميكروبيولوجي",
        "status": "قيد الإعداد — التقديم يبدأ فبراير 2026",
        "verified": (
            "تم التحقق — الموقع الرسمي يعرض المواعيد: Early Bird 8 أبريل، Primary 6 مايو، Final 17 يونيو. "
            "فئة Public Sector & Government Innovation جديدة لعام 2026. "
            "الإعلان عن الفائزين 11 أغسطس. حفل باريس 28 أكتوبر."
        ),
        "notes": (
            "جائزة جديدة غير موجودة في السجل الحالي. "
            "من أعرق الجوائز عالمياً — معترف بها في جميع القطاعات. "
            "القصة: قسم الدراسات نسّق نشر AI عبر كل المختبرات كاستراتيجية موحدة. "
            "يمكن أيضاً التقديم في فئة AI Innovation & Transformation كمشاركة ثانية."
        ),
    },
    {
        "num": 6,
        "section": "قسم الدراسات وتطوير الأعمال",
        "award": "الجائزة الدولية أفكار عربية — Ideas Arabia (19th Cycle)",
        "about": (
            "تهدف إلى التعرف على أهم الأفكار المطبقة للمؤسسات والأفراد. "
            "فئات: الاستدامة، الصحة والسلامة، الابتكار، التقنيات، "
            "إسعاد المتعاملين، المالية، الإنتاجية."
        ),
        "category": "التقنيات",
        "classification": "B (إقليمية/دولية)",
        "website": "https://www.dqg.org/ideas-arabia-international-award",
        "organizer": "مجموعة دبي للجودة (DQG)",
        "deadline": "غير معلن — الدورة 18 انتهت فبراير 2026، الدورة 19 لم تُعلن بعد",
        "fees": "4,450 درهم (تقديم 800 + مشاركة 2,500 + ضريبة)",
        "topic": (
            "الابتكار التنظيمي الذكي: تحول المختبرات المركزية "
            "من الامتثال التقليدي إلى الحوكمة الذكية القائمة على البيانات"
        ),
        "projects": (
            "1. إطار تنظيمي ذكي للمنسوجات — من الامتثال إلى الحوكمة الذكية (الكيميائي)\n"
            "2. التحقق المترولوجي المتنقل OTG لعدادات المحروقات (المقاييس)\n"
            "3. برنامج VeriSoft لرقمنة التحقق القانوني (المقاييس)\n"
            "4. شهادات المطابقة للأبنية الخضراء (منح الشهادات)"
        ),
        "contributing": "الكيميائي، المقاييس، منح الشهادات",
        "status": "انتظار فتح التسجيل للدورة 19",
        "verified": (
            "الجائزة موجودة ومعروفة — مجموعة دبي للجودة. "
            "لكن الدورة 19 لم تُعلن بعد. لا يوجد موعد مؤكد."
        ),
        "notes": (
            "موجودة في السجل — أقسام أخرى تشارك بمواضيع فردية. "
            "قسم الدراسات يقدم محور التحول التنظيمي كقصة موحدة. "
            "يجب متابعة موقع DQG لموعد فتح الدورة 19."
        ),
    },
]

# Write data
for i, p in enumerate(participations):
    row = i + 3
    values = [
        p["num"], p["section"], p["award"], p["about"], p["category"],
        p["classification"], p["website"], p["organizer"], p["deadline"],
        p["fees"], p["topic"], p["projects"], p["contributing"],
        p["status"], p["verified"], p["notes"],
    ]
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.alignment = wrap
        cell.border = thin_border
        if col == 1:
            cell.alignment = Alignment(horizontal="center", vertical="top")

    # Color code verification status column (col 15)
    verify_cell = ws.cell(row=row, column=15)
    if "تم التحقق" in p["verified"]:
        verify_cell.fill = verified_fill
    else:
        verify_cell.fill = unverified_fill

    # Color code urgent deadlines (col 9)
    deadline_cell = ws.cell(row=row, column=9)
    if "09/03" in p["deadline"]:
        deadline_cell.fill = urgent_fill

# Column widths
widths = [4, 22, 28, 40, 25, 12, 35, 25, 18, 12, 40, 50, 25, 20, 40, 40]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

# Row heights
for row in range(3, 9):
    ws.row_dimensions[row].height = 130

output_path = r"C:\Users\LENOVO\Desktop\DCL_Studies_Section_Awards_2026_v3.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
